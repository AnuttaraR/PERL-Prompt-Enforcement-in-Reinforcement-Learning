import json
from openpyxl import Workbook

def process_json_to_sheet(json_data, sheet):
    # Set up the header
    headers = [
        "Questions", "Chunks", "BLEU", "ROUGE1", "ROUGE2", "ROUGEL",
        "METEOR", "Precision", "Recall", "F1", "Chunk Metadata"
    ]
    sheet.append(headers)

    # Iterate over JSON data
    for question, content in json_data.items():
        evaluations = content.get("evaluations", {})
        truncated_texts = content.get("truncated_texts", [])
        first_chunk = True  # Track whether it's the first chunk for the current question

        for chunk, metrics in evaluations.items():
            row = [
                question if first_chunk else "",  # Only add the question for the first chunk
                chunk,
                metrics.get("BLEU", ""),
                str(metrics.get("ROUGE", {}).get("rouge1", "")),  # Convert lists to strings
                str(metrics.get("ROUGE", {}).get("rouge2", "")),
                str(metrics.get("ROUGE", {}).get("rougeL", "")),
                metrics.get("METEOR", ""),
                metrics.get("BERTScore", {}).get("Precision", ""),
                metrics.get("BERTScore", {}).get("Recall", ""),
                metrics.get("BERTScore", {}).get("F1", ""),
                " | ".join([
                    f"{txt['truncated_text']}: {txt['metadata']}" for txt in truncated_texts
                ]) if first_chunk else ""  # Only add metadata for the first chunk
            ]
            sheet.append(row)
            first_chunk = False  # After the first chunk, set this to False


# Load JSON data from files
with open('C:/Users/USER/PycharmProjects/fyp-rnd/evaluation_results_GTE Base.json', 'r') as file1:
    data1 = json.load(file1)

with open('C:/Users/USER/PycharmProjects/fyp-rnd/evaluation_results_BGE Base.json', 'r') as file2:
    data2 = json.load(file2)

with open('C:/Users/USER/PycharmProjects/fyp-rnd/evaluation_results_E5 Base.json', 'r') as file2:
    data3 = json.load(file2)


# Create a new workbook
wb = Workbook()

# Create and populate the first sheet
ws1 = wb.active
ws1.title = "GTE-Base Results"
process_json_to_sheet(data1, ws1)

# Create and populate the second sheet
ws2 = wb.create_sheet(title="BGE-Base Results")
process_json_to_sheet(data2, ws2)

ws3 = wb.create_sheet(title="E5-Base Results")
process_json_to_sheet(data3, ws3)

# Save the workbook
wb.save("D:/output.xlsx")
